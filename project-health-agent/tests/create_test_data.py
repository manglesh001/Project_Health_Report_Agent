"""
Generate test xlsx files that mirror the structure of the actual project plans.
We create two files matching the data from the PDFs.
"""
import openpyxl
from datetime import date, timedelta
from pathlib import Path


def create_unisan_project():
    """create a test xlsx mimicking Project B (UniSan S2P)"""
    wb = openpyxl.Workbook()

    # task sheet
    ws = wb.active
    ws.title = "Project Plan"

    headers = [
        "Level", "Task Name", "Status", "% Complete", "Start Date", "End Date",
        "Duration", "Predecessors", "Owner", "Priority", "Total Float",
        "Critical ?", "On Hold?", "Not Applicable?", "Baseline Start",
        "Baseline Finish", "Variance", "Schedule Health", "Comments", "Assigned To"
    ]
    ws.append(headers)

    # phase 1 tasks (mostly done)
    tasks = [
        [0, "Zycus - UniSan S2P Implementation", "In Progress", 0.44, date(2026,2,11), date(2026,10,9), 170, "", "", "", 0, "TRUE", "", "", date(2026,5,3), date(2026,5,20), -2, "Red", "", ""],
        [1, "Contract Sign Off", "Completed", 1.0, date(2026,2,11), date(2026,2,11), 1, "", "", "", 169, "", "", "", date(2026,2,11), date(2026,2,11), 0, "Green", "", ""],
        [1, "Project Initiation and Data Gathering", "Completed", 1.0, date(2026,2,19), date(2026,3,17), 18, "", "", "", 146, "", "", "", date(2026,2,12), date(2026,3,18), 1, "Green", "", ""],
        [1, "Configuration Validation & Documentation", "Completed", 1.0, date(2026,3,13), date(2026,4,27), 32, "", "", "", 117, "", "", "", date(2026,3,13), date(2026,4,28), 1, "Green", "", ""],
        [1, "Single Sign On and User details Integration", "Completed", 1.0, date(2026,3,13), date(2026,4,16), 25, "", "", "", 124, "", "", "", date(2026,3,13), date(2026,4,16), 0, "Green", "", ""],
        [1, "Contract Management with Basic Setup", "Completed", 1.0, date(2026,4,23), date(2026,5,5), 8, "", "", "", 112, "", "", "", date(2026,4,23), date(2026,5,5), 0, "Green", "", ""],
        [1, "Supplier Information Management with Integration", "Completed", 1.0, date(2026,4,23), date(2026,5,15), 16, "", "", "", 104, "", "", "", date(2026,4,23), date(2026,5,20), 3, "Green", "", ""],
        [1, "User Acceptance Testing (UAT) Phase I", "Completed", 1.0, date(2026,5,4), date(2026,5,20), 13, "", "", "", 101, "", "", "", date(2026,5,4), date(2026,6,16), 18, "Green", "", ""],
        [1, "UAT Completion I", "Completed", 1.0, date(2026,5,20), date(2026,5,20), 0, "107", "", "", 69, "", "", "", date(2026,5,20), date(2026,5,20), 0, "Green", "", ""],
        [1, "Value KPI Sign-off", "Completed", 1.0, date(2026,5,18), date(2026,5,18), 0, "91, 105", "", "", 101, "", "", "", date(2026,5,20), date(2026,5,20), 2, "Red", "", "UniSan Project Group, Zycus Product Specialist, Zycus Project Manager"],
        [1, "Training Phase I", "In Progress", 0.77, date(2026,5,19), date(2026,5,28), 7, "", "", "", 96, "", "", "", date(2026,5,19), date(2026,6,22), 17, "Green", "", ""],
        [1, "Training Completion", "Completed", 1.0, date(2026,5,28), date(2026,5,28), 0, "117", "", "", 96, "", "", "", date(2026,6,22), date(2026,6,22), 17, "Green", "", ""],
        [1, "Production Migration Phase I", "Completed", 1.0, date(2026,5,19), date(2026,6,3), 11, "", "", "", 92, "", "", "", date(2026,5,19), date(2026,6,23), 14, "Green", "", ""],
        [1, "Production Go-Live Phase I", "Completed", 1.0, date(2026,6,8), date(2026,6,8), 0, "146", "", "", 69, "", "", "", date(2026,6,24), date(2026,6,24), 12, "Red", "", ""],
        [1, "Hypercare Phase I", "In Progress", 0.25, date(2026,6,9), date(2026,7,6), 20, "147", "UniSan IT Team, UniSan Integration Team, UniSan Project Group, Zycus Consultant, Zycus Product Specialist", "", 69, "", "", "", date(2026,6,25), date(2026,7,23), 13, "Red", "", ""],

        # phase 2 tasks (trouble area)
        [1, "Configuration Validation & Documentation Phase II", "In Progress", 0.26, date(2026,6,19), date(2026,7,22), 24, "", "", "", 57, "", "", "", None, None, None, "Yellow", "", ""],
        [1, "eProc and eInvoice and Merlin APSD", "In Progress", 0.21, date(2026,6,19), date(2026,8,4), 33, "", "", "", 48, "", "", "", None, None, None, "Green", "", ""],
        [1, "Contract Management with Basic Setup Phase II", "In Progress", 0.03, date(2026,7,14), date(2026,8,7), 19, "", "", "", 45, "", "", "", None, None, None, "Red", "", ""],

        # iSupplier deployment chain (blocked)
        [1, "iSupplier Deployment", "In Progress", 0.13, date(2026,6,12), date(2026,8,11), 43, "", "", "", 43, "", "", "", None, None, None, "Red", "", ""],
        [2, "Design & Development on STAGING Set UP", "In Progress", 0.13, date(2026,6,12), date(2026,8,11), 43, "", "", "", 43, "", "", "", None, None, None, "Green", "", ""],
        [2, "Users,Roles,Workflow & View Definition Configuration", "Not Started", 0.0, date(2026,7,22), date(2026,7,22), 1, "", "", "", 23, "", "", "", None, None, None, "Red", "", ""],
        [2, "Pre-UAT on STAGING Set UP", "In Progress", 0.05, date(2026,6,12), date(2026,7,22), 29, "", "", "", 57, "", "", "", None, None, None, "Green", "", ""],

        # supplier bulk upload chain (cascade blocker)
        [2, "Supplier Bulk Upload on STAGING", "In Progress", 0.05, date(2026,6,12), date(2026,7,15), 24, "", "", "", 62, "", "", "", None, None, None, "Red", "", ""],
        [3, "UniSan to provide supplier bulk upload data", "In Progress", 0.5, date(2026,6,12), date(2026,6,12), 1, "260", "UniSan IT Project Manager, UniSan IT Team, UniSan Project Manager", "", 45, "", "", "", None, None, None, "Red", "", ""],
        [3, "Zycus to review the files and provide feedback", "In Progress", 0.5, date(2026,6,15), date(2026,6,15), 1, "263", "Zycus Product Specialist", "", 45, "", "", "", None, None, None, "Red", "", ""],
        [3, "UniSan to provide updated bulk upload data", "Not Started", 0.0, date(2026,6,23), date(2026,6,24), 2, "264", "UniSan IT Project Manager, UniSan IT Team, UniSan Project Manager", "", 45, "", "", "", None, None, None, "Red", "", ""],
        [3, "Zycus to bulk upload supplier data on STAGING setup", "Not Started", 0.0, date(2026,6,25), date(2026,6,26), 2, "265", "Zycus Product Specialist", "", 45, "", "", "", None, None, None, "Red", "", ""],
        [3, "UniSan to review bulk uploaded data", "Not Started", 0.0, date(2026,6,29), date(2026,7,1), 3, "266", "UniSan Core Project Team", "", 49, "", "", "", None, None, None, "Red", "", ""],
        [3, "UniSan to sign off on supplier bulk upload", "Not Started", 0.0, date(2026,7,1), date(2026,7,1), 1, "267", "UniSan Project Manager", "", 50, "", "", "", None, None, None, "Red", "", ""],
        [3, "UniSan Prod migration Supplier", "Not Started", 0.0, date(2026,7,2), date(2026,7,15), 10, "268", "Zycus Product Specialist", "", 50, "", "", "", None, None, None, "Yellow", "", ""],

        # Integration
        [2, "Integration iSupplier", "In Progress", 0.38, date(2026,6,25), date(2026,7,16), 16, "", "", "", 61, "", "", "", None, None, None, "Green", "", ""],
        [3, "Integration Design", "In Progress", 0.63, date(2026,6,25), date(2026,6,30), 4, "", "", "", 73, "", "", "", None, None, None, "Red", "", ""],
        [4, "Mapping activity for Integration", "Completed", 1.0, date(2026,6,25), date(2026,6,25), 1, "200", "UniSan Integration Team, Zycus Integration Specialist", "", 24, "", "", "", None, None, None, "Green", "", ""],
        [4, "Share TDD for interface", "Completed", 1.0, date(2026,6,26), date(2026,6,26), 1, "273", "Zycus Integration Lead", "", 27, "", "", "", None, None, None, "Green", "", ""],
        [4, "Review of TDD and share feedback with Zycus", "In Progress", 0.5, date(2026,6,29), date(2026,6,29), 1, "274", "UniSan Integration Team", "", 24, "", "", "", None, None, None, "Red", "", ""],
        [4, "Sign off on TDD", "Not Started", 0.0, date(2026,6,30), date(2026,6,30), 1, "275", "UniSan Integration Team", "", 27, "", "", "", None, None, None, "Green", "", ""],
        [3, "Development and Testing on STAGING Set Up", "In Progress", 0.29, date(2026,7,1), date(2026,7,16), 12, "", "", "", 61, "", "", "", None, None, None, "Green", "", ""],
        [4, "Integration set up and development at Zycus", "In Progress", 0.80, date(2026,7,6), date(2026,7,10), 5, "276", "Zycus Integration Specialist", "", 24, "", "", "", None, None, None, "Green", "", ""],
        [4, "Integration development at UniSan - as applicable", "Not Started", 0.0, date(2026,7,1), date(2026,7,7), 5, "276", "UniSan Integration Team", "", 24, "", "", "", None, None, None, "Red", "", ""],
        [4, "Integration Unit Testing and Feedback Incorporation", "Not Started", 0.0, date(2026,7,13), date(2026,7,14), 2, "278, 279", "Zycus Integration Team", "", 24, "", "", "", None, None, None, "Green", "", ""],
        [4, "SIT on STAGING Set Up", "Not Started", 0.0, date(2026,7,15), date(2026,7,16), 2, "280", "UniSan Integration Team, Zycus Integration Team", "", 23, "", "", "", None, None, None, "Green", "", ""],
        [4, "Sign off on SIT", "Not Started", 0.0, date(2026,7,16), date(2026,7,16), 1, "282", "UniSan Integration Team", "", 23, "", "", "", None, None, None, "Green", "", ""],

        # Phase 2 milestones
        [1, "User Acceptance Testing (UAT) Phase II", "Not Started", 0.0, date(2026,8,7), date(2026,8,24), 12, "", "", "", 34, "", "", "", None, None, None, "Green", "", ""],
        [1, "UAT Completion", "Not Started", 0.0, date(2026,8,24), date(2026,8,24), 0, "346, 349", "", "", 24, "", "", "", None, None, None, "Green", "", ""],
        [1, "Training Phase II", "Not Started", 0.0, date(2026,8,18), date(2026,9,2), 12, "", "", "", 27, "", "", "", None, None, None, "Green", "", ""],
        [1, "Production Migration Phase II", "Not Started", 0.0, date(2026,8,21), date(2026,9,7), 12, "", "", "", 24, "", "", "", None, None, None, "Green", "", ""],
        [1, "Production Go-Live Phase II", "Not Started", 0.0, date(2026,9,11), date(2026,9,11), 0, "381", "", "", 0, "", "", "TRUE", None, None, None, "Green", "", ""],
        [1, "Hypercare Phase II", "Not Started", 0.0, date(2026,9,14), date(2026,10,9), 20, "382", "", "", 0, "", "", "TRUE", None, None, None, "Green", "", ""],
    ]

    for task in tasks:
        ws.append(task)

    # summary sheet
    ws2 = wb.create_sheet("Dashboard")
    summary_data = [
        ["Project Name", "", "UniSan S2P Implementation"],
        ["Project Manager", "", "Rajat Bothra"],
        ["Project Start Date", "", date(2026,2,11)],
        ["Project End Date", "", date(2026,10,9)],
        ["Not Started", "", 175],
        ["In Progress", "", 35],
        ["Completed", "", 133],
        ["On Hold", "", 0],
        ["At Risk", "High", ""],
        ["Project Stage", "Training Phase I", ""],
        ["% Complete", "", "44%"],
        ["Schedule Health", "Red", ""],
        ["Today's Date", "", date(2026,7,2)],
    ]
    for row in summary_data:
        ws2.append(row)

    output_path = Path(__file__).parent.parent / "data" / "input" / "Project_B_UniSan.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Created: {output_path}")


def create_titan_project():
    """create a test xlsx mimicking the Titan S2P project"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Plan"

    headers = [
        "Level", "Task Name", "Status", "% Complete", "Start Date", "End Date",
        "Duration", "Predecessors", "Owner", "Priority", "Total Float",
        "Critical ?", "On Hold?", "Not Applicable?", "Baseline Start",
        "Baseline Finish", "Variance", "Schedule Health", "Comments", "Assigned To"
    ]
    ws.append(headers)

    tasks = [
        [0, "Zycus - Titan S2P Implementation", "In Progress", 0.71, date(2025,12,5), date(2026,12,7), 262, "", "", "", 0, "TRUE", "", "", date(2026,5,3), date(2026,5,20), -2, "Green", "", ""],

        # Phase 1 - mostly done
        [1, "Phase 1- S2C", "In Progress", 0.92, date(2025,12,5), date(2026,7,14), 158, "", "abhilasha.s@zycus.com", "", 104, "", "", "", date(2025,12,5), date(2026,7,2), -6, "Yellow", "", ""],

        # Phase 2 - the problem area
        [1, "Phase 2 - P2P", "In Progress", 0.21, date(2026,4,22), date(2026,12,7), 164, "", "", "", 0, "", "", "", date(2026,4,15), date(2026,8,10), -81, "Yellow", "", ""],

        # P2P sub-phases
        [2, "Project Initiation & Data Gathering", "Completed", 1.0, date(2026,4,22), date(2026,6,2), 30, "", "", "", 134, "", "", "", date(2026,4,15), date(2026,4,30), -15, "Yellow", "", ""],
        [3, "Zycus BootCamp", "Completed", 1.0, date(2026,4,22), date(2026,4,23), 2, "", "", "", 162, "", "", "", None, None, None, "Green", "", ""],
        [3, "Logistics Activities", "Completed", 1.0, date(2026,5,18), date(2026,6,2), 12, "", "", "", 134, "", "", "", date(2026,4,15), date(2026,4,20), -9, "Yellow", "", ""],
        [3, "Data Gathering & STAGING Set Up", "Completed", 1.0, date(2026,4,28), date(2026,5,22), 19, "", "", "", 141, "", "", "", date(2026,4,15), date(2026,4,30), -15, "Yellow", "", ""],

        # Configuration Documentation (behind)
        [2, "Configuration Documentation workshop", "In Progress", 0.25, date(2026,6,22), date(2026,7,13), 16, "", "", "", 105, "", "", "", date(2026,4,21), date(2026,5,26), -43, "Yellow", "", ""],
        [3, "P2P, iSupplier Configuration Documentation Workshop", "Completed", 1.0, date(2026,6,22), date(2026,6,26), 5, "", "Titan Core Project Team, Titan Project Manager, Zycus Merlin Product Team, Zycus Project Manager", "", 88, "", "", "", date(2026,4,21), date(2026,5,4), -22, "Yellow", "", ""],
        [3, "Master data to be provided by OTK team", "Not Started", 0.0, date(2026,6,26), date(2026,7,2), 5, "", "Titan Core Project Team", "High", 88, "", "", "", None, None, None, "Red", "", ""],
        [3, "Workflow details & CULT table details to be provided by OTK", "Not Started", 0.0, date(2026,6,26), date(2026,7,10), 11, "106", "Titan Core Project Team", "", 0, "", "", "", None, None, None, "Red", "", ""],
        [3, "Develop & Publish 1st draft of IAD for eProc-eInvoice", "Not Started", 0.0, date(2026,7,7), date(2026,7,9), 3, "292, 293", "Zycus Consultant", "", 88, "", "", "", date(2026,5,12), date(2026,5,12), -24, "Yellow", "", ""],
        [3, "eProc-eInvoice IAD Review and Sign off by Titan", "Not Started", 0.0, date(2026,7,10), date(2026,7,13), 2, "295", "Titan Project Manager", "", 88, "", "", "", date(2026,5,13), date(2026,5,13), -25, "Yellow", "", ""],

        # P2P Integration Design
        [2, "P2P Integration Design", "In Progress", 0.39, date(2026,6,29), date(2026,7,10), 10, "", "", "", 106, "", "", "", None, None, None, "Yellow", "", ""],
        [3, "eInvoice outbound", "In Progress", 0.90, date(2026,6,29), date(2026,6,29), 1, "", "", "", 115, "", "", "", date(2026,5,12), date(2026,5,19), -22, "Yellow", "", ""],
        [3, "eProc Direct & Indirect", "In Progress", 0.90, date(2026,6,30), date(2026,7,1), 2, "", "", "", 113, "", "", "", date(2026,5,12), date(2026,5,18), -30, "Yellow", "", ""],
        [3, "Master data inbound", "Not Started", 0.0, date(2026,7,7), date(2026,7,8), 2, "", "", "", 108, "", "", "", date(2026,5,12), date(2026,5,13), -44, "Yellow", "", ""],
        [3, "Review and feedback incorporation", "Not Started", 0.0, date(2026,7,9), date(2026,7,9), 1, "303", "Titan Integration Team, Titan Project Manager, Zycus Consultant, Zycus Integration Specialist", "", 89, "", "", "", None, None, None, "Yellow", "", ""],
        [3, "Sign off on LLD", "Not Started", 0.0, date(2026,7,10), date(2026,7,10), 1, "304", "Titan Project Manager", "", 89, "", "", "", None, None, None, "Yellow", "", ""],

        # Config & Build
        [2, "Configuration and Build phase", "Not Started", 0.0, date(2026,7,13), date(2026,9,3), 39, "", "", "", 67, "", "", "", None, None, None, "Green", "", ""],
        [3, "User Roles configuration", "Not Started", 0.0, date(2026,7,14), date(2026,7,15), 2, "296", "", "", 88, "", "", "", None, None, None, "Green", "", ""],
        [3, "Admin settings configuration", "Not Started", 0.0, date(2026,7,16), date(2026,7,17), 2, "309", "", "", 88, "", "", "", None, None, None, "Green", "", ""],
        [3, "Configure Workflow", "Not Started", 0.0, date(2026,7,20), date(2026,7,24), 5, "310", "", "", 88, "", "", "", None, None, None, "Green", "", ""],

        # UAT
        [2, "User Acceptance Testing (UAT)", "Not Started", 0.0, date(2026,9,4), date(2026,10,20), 33, "", "", "", 34, "", "", "", None, None, None, "Green", "", ""],
        [3, "Migration of Solution from UAT to PRODUCTION", "Not Started", 0.0, date(2026,9,24), date(2026,10,9), 12, "", "", "", 41, "", "", "", date(2026,7,20), date(2026,8,5), -41, "Yellow", "", ""],
        [3, "Training", "Not Started", 0.0, date(2026,10,12), date(2026,10,20), 7, "", "", "", 34, "", "", "", date(2026,8,3), date(2026,8,7), -38, "Yellow", "", ""],

        # P2P Cutover
        [2, "P2P Cutover", "Not Started", 0.0, date(2026,10,12), date(2026,12,7), 41, "", "", "", 0, "", "", "", date(2026,7,13), date(2026,8,7), -62, "Yellow", "", ""],
        [3, "First Go-Live communication to all suppliers", "Not Started", 0.0, date(2026,10,12), date(2026,10,13), 2, "477", "Titan Project Team", "", 0, "", "", "TRUE", date(2026,7,13), date(2026,7,13), -59, "Yellow", "", ""],
        [3, "Close all requisitions/PO/invoices in existing system", "Not Started", 0.0, date(2026,10,16), date(2026,10,22), 5, "485", "Titan Project Team", "", 0, "", "", "TRUE", date(2026,8,6), date(2026,8,6), -50, "Yellow", "", ""],
        [3, "Load delta master data via integration", "Not Started", 0.0, date(2026,10,30), date(2026,11,5), 5, "487", "Titan Integration Team", "", 0, "", "", "TRUE", date(2026,7,31), date(2026,8,3), -63, "Yellow", "", ""],
        [3, "Go Live", "Not Started", 0.0, date(2026,11,9), date(2026,11,9), 1, "", "", "", 20, "", "", "", date(2026,8,10), date(2026,8,10), -61, "Yellow", "", ""],
    ]

    for task in tasks:
        ws.append(task)

    # summary
    ws2 = wb.create_sheet("Dashboard")
    summary_data = [
        ["Project Name", "", "Titan S2P Implementation"],
        ["Project Manager", "", "Aftab Hashambhai"],
        ["Project Start Date", "", date(2025,12,5)],
        ["Project End Date", "", date(2026,12,7)],
        ["Not Started", "", 194],
        ["In Progress", "", 15],
        ["Completed", "", 255],
        ["On Hold", "", 3],
        ["At Risk", "High", ""],
        ["Project Stage", "Configuration and Build phase", ""],
        ["% Complete", "", "71%"],
        ["Schedule Health", "Green", ""],
        ["Today's Date", "", date(2026,7,2)],
    ]
    for row in summary_data:
        ws2.append(row)

    output_path = Path(__file__).parent.parent / "data" / "input" / "S2P_Project_Titan.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    print(f"Created: {output_path}")


if __name__ == "__main__":
    create_unisan_project()
    create_titan_project()
    print("\nDone. Test files created in data/input/")
